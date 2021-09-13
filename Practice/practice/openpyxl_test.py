from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt

#read file
wb = load_workbook("/home/fanelesibonge/Documents/Excel/test.xlsx",data_only=True)

#access specific sheet
ws = wb['Normalize']

data_rows = []

for row in ws['B5':'D527']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

df = pd.DataFrame(data_rows).dropna()
df = df.rename(columns=df.iloc[0]).drop(df.index[0])


print(list(df.columns))
'''
#df.plot(x='Date')
#plt.show()

data = []
started = False
#select column and print it
for cell in ws['B']:
    if cell.value == None:
        if started == True:
            print("I am none")
            break
    elif cell.value != None:
        started = True
    
        data.append(cell.value)

print(data)

print(f"Rows: {len(list(ws.rows))}")'''

def read_columns(ws, begin, columns, end=0):
    data_rows = []
    df = None
    if end == 0:
        #return [ws["{}{}".format(column, row)].value for row in range(begin,len(list(ws.rows))+1) for column in columns]
        for row in range(begin,len(list(ws.rows))+1):
            data_cols = []
            for column in columns:
                data_cols.append(ws["{}{}".format(column,row)].value)
            data_rows.append(data_cols)
    elif end > 0:
        for row in range(begin,end+1):
            data_cols = []
            for column in columns:
                data_cols.append(ws["{}{}".format(column,row)].value)
            data_rows.append(data_cols)
    
    df = pd.DataFrame(data_rows).dropna()
    
    print(list(df.columns))
    return df

print(df.iloc[0])


df = read_columns(ws, 5,'BEFGH')
df = df.rename(columns=df.iloc[0]).drop(df.index[0])
print(list(df.columns))

df.plot(x='Date/Symbol')
plt.show()

fil = open('Practice/misc/openpyxl.txt','w')
df.to_string(fil)
fil.close()


print(df)
print(len('BEF'))