import tkinter as tk
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import tkinter.messagebox as mb

from scrollable_frame import ScrollableFrame

class SelectCellsFrame(tk.Frame):
    
    def __init__(self, master):
        tk.Frame.__init__(self, master)
        self.frame = ScrollableFrame(master)

        print(master.mChosenSheet.name)
        
        workbook = openpyxl.load_workbook(master.mFilePath)
        
        self.ws = workbook[master.mChosenSheet.name]
        self.buildWidgets(master)
        self.frame.pack()

    def __call__(self, master):
        SelectCellsFrame(master).pack(fill='x',expand='yes')

    def buildWidgets(self, master):

        big_row = tk.Frame(self.frame.scrollable_frame)
        row = tk.Frame(big_row)
        lab = tk.Label(row, text='Select Sheet cells to be plotted',font=('Arial', 17, 'bold'))
        row.pack(side='top',fill='x')
        lab.pack(side='top')
        
        row = tk.Frame(big_row)
        hint = tk.Label(row, text='Select one of the two options ')
        row.pack(side='top')
        hint.pack(side='top')
        
        big_row.pack(side='top')

        self.var = tk.StringVar()           #variable for radio button check
        row = tk.Frame(self.frame.scrollable_frame)
        lab = tk.Label(row, text='Select One table',font=('Arial',17,'bold'))
        hint = tk.Label(row, text='(enter the top most left cell and the bottom most right)',font=('Arial',12,'italic'))
        row.pack(side='top',expand='yes',pady=20)
        tk.Radiobutton(row, text='',value='Table',variable=self.var).pack(side='left')
        lab.pack(side='top')
        hint.pack(side='top')

        #top left cell
        row = tk.Frame(self.frame.scrollable_frame)
        lab = tk.Label(row, text='Top Left Cell (i.e. A10)', font=('Arial', 15, 'bold'),relief='ridge',width='30')
        self.top_left_cell = tk.Entry(row,font=('Arial',15,'bold'),width=20,relief='sunken')
        row.pack(side='top',padx=55,pady=10)
        lab.pack(side='left')
        self.top_left_cell.pack(side='right')

        #bottom right cell
        row = tk.Frame(self.frame.scrollable_frame)
        lab = tk.Label(row, text='Bottom Right Cell (i.e. F43)',font=('Arial', 15, 'bold'),relief='ridge',width='30')
        self.bottom_right_cell = tk.Entry(row, font=('Arial',15,'bold'),width=20,relief='sunken')
        row.pack(side='top',pady=10)
        lab.pack(side='left')
        self.bottom_right_cell.pack(side='right')

    
        #multiple column selection definition
        row = tk.Frame(self.frame.scrollable_frame)
        lab = tk.Label(row, text='Select multiple columns to plot',font=('Arial',17,'bold'))
        #hint = tk.Label(row, text='(choose top and bottom cells for the desired column. The x-axis variable \nshould be specified first)',font=('Arial',12,'italic'))
        row.pack(side='top',pady=30)
        tk.Radiobutton(row, text='',value='Columns',variable=self.var).pack(side='left')
        lab.pack(side='top')
        #hint.pack(side='top')

        #columns for multiple selection
        big_row = tk.Frame(self.frame.scrollable_frame)
        row = tk.Frame(big_row)
        lab = tk.Label(row, text='Columns',font=('Arial',15,'bold'),width=30, relief='ridge')
        self.multiple_columns = tk.Entry(row,font=('Arial',15,'bold'),width=20,relief='sunken')
        row.pack(side='top')
        lab.pack(side='left')
        self.multiple_columns.pack(side='right')
        
        row = tk.Frame(big_row)
        hint = tk.Label(row, text='(List the letters of the columns desired (i.e. ABEF))',font=('Arial',12,'italic'))
        row.pack(side='top')
        hint.pack(side='top')
        big_row.pack(side='top')
        
        
        #start row number
        row = tk.Frame(self.frame.scrollable_frame)
        lab = tk.Label(row, text='Start Row Number',font=('Arial',15, 'bold'),relief='ridge',width=30)
        self.start_row = tk.Entry(row,width=20,font=('Arial',15,'bold'),relief='sunken')
        row.pack(side='top')
        lab.pack(side='left')
        self.start_row.pack(side='top')
        
        #end row number
        big_row = tk.Frame(self.frame.scrollable_frame)
        row = tk.Frame(big_row)
        lab = tk.Label(row, text='End Row Number',font=('Arial',15,'bold'),relief='ridge',width=30)
        self.end_row = tk.Entry(row,width=20,font=('Arial',15,'bold'),textvariable=tk.StringVar(row,value='0'),relief='sunken')
        row.pack(side='top')
        lab.pack(side='left')
        self.end_row.pack(side='right')
        
        row = tk.Frame(big_row)
        hint = tk.Label(row, text='(If empty, data will be read until the end of the column)',font=('Arial',12,'italic'))
        row.pack(side='top')
        hint.pack(side='top')
        big_row.pack(side='top',pady=20)
        
        #compulsory section
        row = tk.Frame(self.frame.scrollable_frame)
        lab = tk.Label(row, text='Fill In Below:',font=('Arial',17,'bold'))
        row.pack(side='top',pady=20)
        lab.pack(side='top')
        
        #x-axis name/ label
        row = tk.Frame(self.frame.scrollable_frame)
        lab = tk.Label(row,text='x-axis label',font=('Arial',15,'bold'),width=30,relief='ridge')
        self.x_axis_label = tk.Entry(row, width=20,font=('Arial',15,'bold'),relief='sunken')
        row.pack(side='top',pady=15)
        lab.pack(side='left')
        self.x_axis_label.pack(side='right')
        
        #y-axis label
        row = tk.Frame(self.frame.scrollable_frame)
        lab = tk.Label(row, text='y-axis label',font=('Arial',15,'bold'),width=30,relief='ridge')
        self.y_axis_label = tk.Entry(row,width=20,font=('Arial',15,'bold'),relief='sunken')
        row.pack(side='top',pady=15)
        lab.pack(side='left')
        self.y_axis_label.pack(side='right')
        
        #name of x axis column
        row = tk.Frame(self.frame.scrollable_frame)
        lab = tk.Label(row, text='x-axis column name',font=('Arial',15,'bold'),width=30,relief='ridge')
        self.x_name = tk.Entry(row, width=20,font=('Arial',15,'bold'),relief='sunken')
        row.pack(side='top')
        lab.pack(side='left')
        self.x_name.pack(side='right')
        
        #button to plot selected
        row = tk.Frame(self.frame.scrollable_frame)
        but = tk.Button(row,text='Plot',font=('Arial',15,'bold'),width=20,command=lambda:self.plotGraph())
        row.pack(side='top', pady=15)
        but.pack(side='top')
        
        #button to select new sheet
        row = tk.Frame(self.frame.scrollable_frame)
        but = tk.Button(row, text='Select New Sheet',font=('Arial',15,'bold'),width=20,command=lambda:self.selectSheet(master))
        row.pack(side='top',pady=15)
        but.pack(side='top')
        
        #button to select a different file
        row = tk.Frame(self.frame.scrollable_frame)
        but = tk.Button(row, text='Select New File',font=('Arial',15,'bold'),width=20,command=lambda:self.selectFile(master))
        row.pack(side='top',pady=15)
        but.pack(side='top')
    
    #function to handle selecting a new sheet
    def selectSheet(self,master):        
        from select_sheet_frame import SelectSheet
        self.frame.destroyMe()
        master.switch_frame(SelectSheet)
        
    #function to select new file
    def selectFile(self, master):
        from select_file_frame import SelectFileFrame
        self.frame.destroyMe()
        master.switch_frame(SelectFileFrame)
    
    #function to read columns and draw graphs
    def plotGraph(self):
        data_row = []
        df = None
        
        if self.var.get() == 'Table':
            print('I am in Table selection')
            print(f'Worksheet {self.ws}')
            for row in self.ws[f'{self.top_left_cell.get()}':f'{self.bottom_right_cell.get()}']:
                data_cols = []
                for cell in row:
                    data_cols.append(cell.value)
                data_row.append(data_cols)
                
        elif self.var.get() == 'Columns': 
            print('I am in columns selection')
            if int(self.end_row.get()) == 0:
                print('I am in end == 0')
                for row in range(int(self.start_row.get()), len(list(self.ws.rows))+1):
                    data_cols = []
                    for column in self.multiple_columns.get():
                        data_cols.append(self.ws["{}{}".format(column,row)].value)
                    data_row.append(data_cols)
            elif int(self.end_row.get()) > 0:
                print('I am in end > 0')
                for row in range(int(self.start_row.get()),int(self.end_row.get())+1):
                    data_cols = []
                    for column in self.multiple_columns.get():
                        data_cols.append(self.ws["{}{}".format(column,row)].value)
                    data_row.append(data_cols)
        else:
            print('I am in no selection')
            mb.showinfo('Notice', 'Please choose one of the two options: Select One Table or Select multiple columns')
            return
        df = pd.DataFrame(data_row)
        df = df.rename(columns=df.iloc[0]).drop(df.index[0])
        
        df.plot(x=self.x_name.get())
        plt.xlabel(self.x_axis_label.get())
        plt.ylabel(self.y_axis_label.get())
        plt.show()
        
        print(df)